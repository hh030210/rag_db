# -*- coding: utf-8 -*-
"""
unified_ingest.py
=================
将 data_input/insert_Data 和 data_input/test_data 的全部文件
分片入库到一个新的 Qdrant collection。

不依赖 LLM chunker，使用纯文本固定窗口分块 + BGE 向量编码。

使用方式:
    python unified_ingest.py --collection unified_corpus
"""
import argparse
import json
import math
import re
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

# ── 配置 ───────────────────────────────────────────────────────────
DEFAULT_COLLECTION = "unified_corpus"
DEFAULT_CHUNK_SIZE = 600       # 字符数（含 overlap）
DEFAULT_OVERLAP = 80           # 重叠字符数
BATCH_SIZE = 50                # 每批写入 Qdrant 的点数
QDRANT_HOST = "127.0.0.1"
QDRANT_PORT = 6333
# ───────────────────────────────────────────────────────────────────

import httpx

PROJECT_ROOT = Path(__file__).parent
sys.path.insert(0, str(PROJECT_ROOT))
from retrieval_fusion_eval import _load_bge_encoder, _encode_query


# ══════════════════════════════════════════════════════════════════
# 文本分块工具
# ══════════════════════════════════════════════════════════════════

def chunk_text(text: str, chunk_size: int = 600, overlap: int = 80) -> list[str]:
    """按字符数滑动窗口分块，保留段落边界。"""
    # 先按段落分割
    paras = re.split(r"\n{2,}", text.strip())
    chunks = []
    buf = ""
    for para in paras:
        para = para.strip()
        if not para:
            continue
        if len(buf) + len(para) + 1 <= chunk_size:
            buf = (buf + "\n" + para).strip()
        else:
            if buf:
                chunks.append(buf)
            # 如果段落本身超长，按句子继续切
            while len(para) > chunk_size:
                # 按句子或逗号分割
                sentences = re.split(r"(?<=[。！？；\n])", para)
                cur = ""
                for s in sentences:
                    if len(cur) + len(s) <= chunk_size:
                        cur += s
                    else:
                        if cur:
                            chunks.append(cur.strip())
                        cur = s
                if cur.strip():
                    chunks.append(cur.strip())
                break
            else:
                buf = para
    if buf.strip():
        chunks.append(buf.strip())
    return chunks


def extract_xlsx_text(file_path: Path) -> list[dict]:
    """从 xlsx 提取 query/answer 对"""
    try:
        import openpyxl
    except ImportError:
        return []
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
    except Exception:
        return []
    records = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        q_idx = next((i for i, h in enumerate(headers) if "query" in h or "问题" in h), 0)
        a_idx = next((i for i, h in enumerate(headers) if "answer" in h or "答案" in h), -1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            q = str(row[q_idx]).strip() if row[q_idx] else ""
            a = str(row[a_idx]).strip() if a_idx >= 0 and row[a_idx] else ""
            if q:
                # 把 Q 和 A 合并成一个 chunk
                records.append({"chunk_text": f"问题：{q}\n答案：{a}" if a else f"问题：{q}\n答案：暂无", "genre": "多跳问答"})
    return records


# ══════════════════════════════════════════════════════════════════
# Qdrant 操作
# ══════════════════════════════════════════════════════════════════

def api(url: str, method="GET", json_body=None, params=None):
    client = httpx.Client(base_url=f"http://{QDRANT_HOST}:{QDRANT_PORT}", timeout=60)
    kw = {"json": json_body} if json_body is not None else {}
    if params:
        kw["params"] = params
    r = client.request(method, url, **kw)
    return r


def create_collection(name: str, vector_size: int = 1024):
    r = api(f"/collections/{name}")
    if r.status_code == 200:
        print(f"  Collection '{name}' 已存在，跳过创建。")
        return
    # 创建新 collection: PUT /collections/{name} + vectors 配置
    body = {
        "vectors": {
            "chunk_text_vec": {
                "size": vector_size,
                "distance": "Cosine"
            }
        }
    }
    r = api(f"/collections/{name}", method="PUT", json_body=body)
    if r.status_code in (200, 201):
        print(f"  Collection '{name}' 创建成功。")
    else:
        print(f"  创建 collection 失败: {r.status_code} {r.text[:200]}")


def delete_collection(name: str):
    r = api(f"/collections/{name}", method="DELETE")
    print(f"  删除 collection '{name}': {r.status_code}")


def get_collection_count(name: str) -> int:
    r = api(f"/collections/{name}")
    if r.status_code == 200:
        return r.json().get("result", {}).get("points_count", 0)
    return 0


def upsert_batch(name: str, points: list) -> bool:
    r = api(f"/collections/{name}/points", method="PUT", json_body={"points": points})
    if r.status_code not in (200, 201):
        print(f"  [错误] upsert: {r.status_code} {r.text[:300]}")
        return False
    return True


# ══════════════════════════════════════════════════════════════════
# 文件读取
# ══════════════════════════════════════════════════════════════════

def read_file(file_path: Path) -> list[dict]:
    """根据文件类型读取并分块，返回 [{chunk_text, genre, doc_title}, ...]"""
    suffix = file_path.suffix.lower()
    name = file_path.stem
    parent = file_path.parent.name

    if suffix == ".txt":
        text = file_path.read_text(encoding="utf-8")
        chunks = chunk_text(text, DEFAULT_CHUNK_SIZE, DEFAULT_OVERLAP)
        genre = "南孔文化"
        return [{"chunk_text": c, "genre": genre, "doc_title": file_path.name} for c in chunks]

    elif suffix == ".md":
        text = file_path.read_text(encoding="utf-8")
        # 判定景区名
        spot = _guess_spot(name)
        genre = _guess_genre(name, spot, parent)
        chunks = chunk_text(text, DEFAULT_CHUNK_SIZE, DEFAULT_OVERLAP)
        return [{"chunk_text": c, "genre": genre, "doc_title": file_path.name,
                 "spot": spot} for c in chunks]

    elif suffix == ".xlsx":
        records = extract_xlsx_text(file_path)
        return records

    return []


def _guess_spot(name: str) -> str:
    """从文件名猜测景区名"""
    known = ["南孔庙", "少林寺", "张家界", "明十三陵", "西湖", "颐和园", "龙门石窟",
             "南孔文化", "西双版纳", "多景区"]
    for s in known:
        if s in name:
            return s
    return "通用"


def _guess_genre(name: str, spot: str, parent: str) -> str:
    """猜测 genre"""
    if "景区介绍" in name or "景点介绍" in name:
        return "景区介绍"
    if "运营信息" in name or "运营介绍" in name:
        return "运营信息"
    if parent == "insert_Data":
        if "多景区" in name or "知识" in name:
            return "景区知识"
        if "多跳" in name:
            return "多跳问答"
        if "南孔" in name:
            return "南孔文化"
    return "通用"


# ══════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--collection", default=DEFAULT_COLLECTION)
    parser.add_argument("--recreate", action="store_true", help="删除旧 collection 后重建")
    args = parser.parse_args()

    coll_name = args.collection
    print(f"\n{'='*60}")
    print(f"  统一入库流程")
    print(f"  Collection: {coll_name}")
    print(f"  数据来源: data_input/insert_Data + data_input/test_data")
    print(f"{'='*60}\n")

    # ── Step 1: 准备输出目录 ─────────────────────────────────────
    OUTPUT_DIR = PROJECT_ROOT / "output_chunks_unified"
    OUTPUT_DIR.mkdir(exist_ok=True)

    # ── Step 2: 加载 encoder ────────────────────────────────────
    print("[1/5] 加载 BGE encoder...")
    encoder = _load_bge_encoder()
    print("  BGE encoder 加载完成。")

    # ── Step 3: 收集并处理所有文件 ─────────────────────────────
    print("\n[2/5] 扫描并读取文件...")
    all_chunks: list[dict] = []
    stats = {}  # doc_title -> count

    data_dirs = [
        PROJECT_ROOT / "data_input" / "insert_Data",
        PROJECT_ROOT / "data_input" / "test_data",
    ]

    for data_dir in data_dirs:
        if not data_dir.exists():
            print(f"  目录不存在: {data_dir}")
            continue
        for file_path in sorted(data_dir.iterdir()):
            if not file_path.is_file():
                continue
            suffix = file_path.suffix.lower()
            if suffix not in (".txt", ".md", ".xlsx"):
                continue
            print(f"  处理: [{data_dir.name}] {file_path.name}")
            items = read_file(file_path)
            # 写入 chunks JSON 供检查
            json_out = OUTPUT_DIR / f"{file_path.stem}_chunks.json"
            json_out.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
            for item in items:
                all_chunks.append({
                    "source": file_path.name,
                    **item
                })
            stats[file_path.name] = len(items)
            print(f"    -> {len(items)} chunks")

    print(f"\n  总计: {len(all_chunks)} chunks（来自 {len(stats)} 个文件）")
    for fname, cnt in sorted(stats.items()):
        print(f"    {fname}: {cnt}")

    # ── Step 4: 创建 / 清空 collection ────────────────────────
    print(f"\n[3/5] {'创建' if args.recreate else '使用已有'} collection '{coll_name}'...")
    if args.recreate:
        delete_collection(coll_name)
    create_collection(coll_name)

    current_count = get_collection_count(coll_name)
    print(f"  当前 collection 点数: {current_count}")
    start_id = current_count + 1

    # ── Step 5: 批量写入 Qdrant ─────────────────────────────────
    print(f"\n[4/5] 写入 Qdrant (批量={BATCH_SIZE})...")
    total_written = 0
    batch = []
    t0 = time.time()

    for i, chunk in enumerate(all_chunks):
        vec = _encode_query(encoder, chunk["chunk_text"])
        if vec is None:
            print(f"  [跳过] 向量编码失败: {chunk.get('source','?')}")
            continue

        pt = {
            "id": start_id + i,
            "vector": {"chunk_text_vec": vec},
            "payload": {
                "chunk_id": f"{Path(chunk['source']).stem}_c{i:04d}",
                "doc_id": Path(chunk["source"]).stem,
                "doc_title": chunk.get("doc_title", chunk["source"]),
                "genre": chunk.get("genre", "通用"),
                "chunk_text": chunk["chunk_text"],
                "chunk_text_full": chunk["chunk_text"],
                "spot": chunk.get("spot", ""),
            }
        }
        batch.append(pt)

        if len(batch) >= BATCH_SIZE:
            ok = upsert_batch(coll_name, batch)
            if ok:
                total_written += len(batch)
                pct = total_written * 100 // len(all_chunks)
                print(f"  写入 {total_written}/{len(all_chunks)} ({pct}%)")
            batch = []

    if batch:
        ok = upsert_batch(coll_name, batch)
        if ok:
            total_written += len(batch)
            print(f"  写入 {total_written}/{len(all_chunks)} (100%)")

    elapsed = time.time() - t0

    # ── Step 6: 验证 ───────────────────────────────────────────
    print(f"\n[5/5] 验证入库结果...")
    new_count = get_collection_count(coll_name)
    print(f"\n{'='*50}")
    print(f"入库完成! 新增 {total_written} chunks")
    print(f"耗时: {elapsed:.1f}s")
    print(f"'{coll_name}' 总点数: {new_count} (之前 {current_count})")

    # 按 genre 统计
    offset = None
    genre_counts = {}
    checked = 0
    while checked < new_count:
        body = {"limit": 500, "with_payload": True, "with_vector": False}
        if offset:
            body["offset"] = offset
        r = api(f"/collections/{coll_name}/points/scroll", method="POST", json_body=body)
        pts = r.json().get("result", {}).get("points", [])
        if not pts:
            break
        for p in pts:
            g = (p.get("payload") or {}).get("genre", "未知")
            genre_counts[g] = genre_counts.get(g, 0) + 1
        checked += len(pts)
        offset = r.json().get("result", {}).get("next_page_offset")
        if not offset:
            break

    print(f"\n按 genre 统计:")
    for g, n in sorted(genre_counts.items(), key=lambda x: -x[1]):
        print(f"  {g}: {n}")


if __name__ == "__main__":
    main()
